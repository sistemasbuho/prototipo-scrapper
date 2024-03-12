import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill


workbook = openpyxl.Workbook()
sheet = workbook.active



# def generador_excel(lista_scraper: list):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active

#     columnas = list(lista_scraper[0].keys())
#     for col_idx, columna in enumerate(columnas, start=1):
#         sheet.cell(row=1, column=col_idx, value=columna)

#     for row_idx, objeto in enumerate(lista_scraper, start=2):
#         for col_idx, columna in enumerate(columnas, start=1):
#             sheet.cell(row=row_idx, column=col_idx, value=objeto[columna])

#     return workbook


# Excel con mapeo de columnas
def generador_excel(lista_scraper: list, column_mapping: dict, order_list: list):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Obtener las columnas según el mapeo proporcionado
    columnas = list(column_mapping.keys())
    for col_idx, columna in enumerate(columnas, start=1):
        sheet.cell(row=1, column=col_idx, value=columna)
        
    # Ordenar las filas según la lista de orden
    lista_scraper_ordenada = sorted(lista_scraper, key=lambda x: order_list.index(x["url"]))

    # Llenar los datos según el mapeo
    for row_idx, objeto in enumerate(lista_scraper_ordenada, start=2):
        for col_idx, columna in enumerate(columnas, start=1):
            ### Obtener la clave correspondiente según el mapeo
            clave = column_mapping[columna]
            ### Escribir el valor en la celda
            sheet.cell(row=row_idx, column=col_idx, value=objeto[clave])
            
    # Colorear las filas donde el contenido es None o vacío de rojo
    red_fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
    orange_fill = PatternFill(start_color='FFFFA500',end_color='FFFFA500',fill_type='solid')
    for row in range(2, len(lista_scraper) + 2):
        content_value = sheet.cell(row=row, column=columnas.index("Contenido") + 1).value
        author_value = sheet.cell(row=row, column=columnas.index("author") + 1).value
        if len(content_value) < 1000 or (author_value is None or author_value == '' or author_value=="None\n"):
            for col in range(1, len(columnas) + 1):
                sheet.cell(row=row, column=col).fill = orange_fill
        if content_value is None or content_value == '' or content_value=="None\n" or content_value=="Invalid URL" or len(content_value) < 100:
            for col in range(1, len(columnas) + 1):
                sheet.cell(row=row, column=col).fill = red_fill

    return workbook

def descargar_excel(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer