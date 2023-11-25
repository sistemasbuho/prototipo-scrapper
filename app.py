from flask import Flask, render_template, request, redirect, url_for, send_from_directory, current_app, jsonify
import openpyxl
import requests
import logging
from celery import Celery
import os
import time

app = Flask(__name__)

logger = logging.getLogger(__name__)

# Configuración de Celery
app.config['CELERY_BROKER_URL'] = 'redis://prototipo-scrapper_redis_1:6379/0'
app.config['CELERY_RESULT_BACKEND'] = 'redis://prototipo-scrapper_redis_1:6379/0'
celery = Celery(app.name, broker=app.config['CELERY_BROKER_URL'])

# Configura SERVER_NAME para que coincida con la forma en que se ejecuta tu aplicación
app.config['SERVER_NAME'] = 'pruebas-scraper.buho.media'  # Ajusta el puerto según tu configuración


celery.conf.update(app.config)

def unir_descripcion_content(response:dict):
    response["content"] = f"{response['og_description']}\n\n{response['content']}"
    return response


def limpieza(response):
    response = response.json()
    del response["effective_url"]
    del response["word_count"]
    del response["og_url"]
    del response["og_image"]
    del response["og_type"]
    del response["twitter_site"]
    del response["twitter_creator"]
    del response["twitter_image"]
    del response["twitter_title"]
    del response["twitter_description"]
    del response["twitter_card"]

    del response["title"]
    del response["excerpt"]
    del response["date"]
    del response["author"]
    del response["language"]
    del response["og_title"]
    response = unir_descripcion_content(response)
    del response["og_description"]
    return response

def txtfly(url):
    api = "https://full-text-rss.p.rapidapi.com/extract.php"
    payload = {
        "url": f"{url}",
        "xss": "1",
        "lang": "2",
        "links": "remove",
        "content": "text"
    }
    headers = {
        "content-type": "application/x-www-form-urlencoded",
        "X-RapidAPI-Key": "6fd32a4cc6mshb266cfc4f75cd1dp12d4d1jsnd5bc578c4926",
    }
    data = requests.post(api, data=payload, headers=headers)
    data_limpia = limpieza(data)
    return data_limpia

@celery.task
def scrape_task(urls):
    # logger.info("Entro a la tarea#10")
    lista_scraper = []
    for url in urls:
        lista_scraper.append(txtfly(url))

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    columnas = list(lista_scraper[0].keys())
    for col_idx, columna in enumerate(columnas, start=1):
        sheet.cell(row=1, column=col_idx, value=columna)

    for row_idx, objeto in enumerate(lista_scraper, start=2):
        for col_idx, columna in enumerate(columnas, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=objeto[columna])
    # Guardar el archivo con los resultados
    filename = "scrapers.xlsx"
    workbook.save(filename)

    return filename  # Retorna el nombre del archivo


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scrape', methods=['POST'])
def scrape():
    if 'file' not in request.files:
        return jsonify({'message': 'No se ha proporcionado un archivo para el scraper'})

    file = request.files['file']

    if file.filename == '':
        return jsonify({'message': 'Nombre de archivo vacío'})

    if file:
        file.save('urls.xlsx')

        # Leer las URLs desde el archivo Excel
        workbook = openpyxl.load_workbook('urls.xlsx')
        sheet = workbook.active
        urls = [cell.value for cell in sheet['A']]

        scrape_task.apply_async(args=(urls,))
        return jsonify({'message': 'La tarea de raspado se está ejecutando en segundo plano. Puede tomar un tiempo.'})

# Ruta para verificar la existencia del archivo "scraper.xlsx"
@app.route('/check_scraper_existence')
def check_scraper_existence():
    scraper_exists = os.path.exists('scrapers.xlsx')
    return jsonify({'exists': scraper_exists})


@app.route('/download_scraper')
def download_scraper():
    directory = os.getcwd()
    filename = "scrapers.xlsx"

    try:
        file_path = os.path.join(directory, filename)
        if os.path.exists(file_path):
            return send_from_directory(directory, filename, as_attachment=True)
    finally:
        file_path = os.path.join(directory, filename)
        if os.path.exists(file_path):
            os.remove(file_path)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050)